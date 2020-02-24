#!/usr/bin/env python
# -*- coding: utf-8 -*- #
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from openpyxl import load_workbook
import os 
import time
import ConfigParser

dirpath = os.getcwd()
chromePath = dirpath+'\chromedriver\chromedriver.exe'
phantom = dirpath+'\\phantomjs\\bin\\phantomjs.exe'
phantomjsPath = str.replace(phantom,'\\','\\\\')

config = ConfigParser.RawConfigParser()
config.read(dirpath+'\ConfigFile.properties')

def huayClub():

    huayClub_url = config.get('HuayClubSection', 'huayClub_url')
    huayClub_user = config.get('HuayClubSection', 'huayClub_user')
    huayClub_pass = config.get('HuayClubSection', 'huayClub_pass')
    try: 
        driver = webdriver.PhantomJS(executable_path = phantomjsPath)
        driver.maximize_window()
        driver.get(huayClub_url)
        print driver.current_url
        print '=============================='
        print '>>> HuayClub <<<<'

        elem = driver.find_element_by_name('username')
        elem.send_keys(huayClub_user)
        elem = driver.find_element_by_name('password')
        elem.send_keys(huayClub_pass)
        elem = WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="login-box"]/div/div/form/fieldset/div[2]/button')))
        elem.click()
        elem = WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="sidebar-menu"]/li[1]/a')))
        elem.click()
        elem = WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="sidebar-menu"]/li[1]/ul/li[2]/a')))
        elem.click()
        elem = WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.ID, 'week')))
        elem.click()
        elem = driver.find_element_by_xpath('//*[@id="game-type-list"]/div[2]/button')
        elem.click()

        sumAgent =  WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="by-member-table-2"]/tbody/tr/td[11]'))).text
        sumCom = WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="by-member-table-2"]/tbody/tr/td[15]'))).text

        print'Sum Agent = ' ,sumAgent
        print 'Sum Company = ' ,sumCom

        wb = load_workbook(dirpath+'\output\Report.xlsx')
        print '[HuayClub] Opening Excel File..'
        ws = wb['Sheet1']
        ws['B19'].value=sumAgent
        ws['H19'].value=sumCom
        wb.save(dirpath+'\output\Report.xlsx')
        print '[HuayClub] Save Excel Success!!'
        print '=============================='
    except:
        print "Something Error"


def huayClub2():
    try: 
        huayClub_url = config.get('HuayClubSection', 'huayClub_url')
        huayClub_user = config.get('HuayClubSection', 'huayClub_user')
        huayClub_pass = config.get('HuayClubSection', 'huayClub_pass')

        huayClub_driver = webdriver.Chrome(chromePath)
        huayClub_driver.maximize_window()
        huayClub_driver.set_page_load_timeout(10)
        
        print 'HuayClub'
    
        huayClub_driver.get('https://agent.superlot999.com/login')
        elem = huayClub_driver.find_element_by_name('username')
        elem.send_keys(huayClub_user)
        elem = huayClub_driver.find_element_by_name('password')
        elem.send_keys(huayClub_pass)
        elem = WebDriverWait(huayClub_driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="login-box"]/div/div/form/fieldset/div[2]/button')))
        elem.click()
        elem = WebDriverWait(huayClub_driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="sidebar-menu"]/li[1]/a')))
        elem.click()
        elem = WebDriverWait(huayClub_driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="sidebar-menu"]/li[1]/ul/li[2]/a')))
        elem.click()
        elem = WebDriverWait(huayClub_driver, 10).until(ec.visibility_of_element_located((By.ID, 'week')))
        elem.click()
        elem = huayClub_driver.find_element_by_xpath('//*[@id="game-type-list"]/div[2]/button')
        elem.click()

        sumAgent =  WebDriverWait(huayClub_driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="by-member-table-2"]/tbody/tr/td[11]'))).text
        sumCom = WebDriverWait(huayClub_driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="by-member-table-2"]/tbody/tr/td[15]'))).text
        
        print 'Sum Agent = ' ,sumAgent
        print 'Sum Company = ' ,sumCom

        wb = load_workbook(dirpath+'\output\Report.xlsx')
        print '[HuayClub] Opening Excel File..'
        ws = wb['Sheet1']
        ws['B19'].value=sumAgent
        ws['H19'].value=sumCom
        wb.save(dirpath+'\output\Report.xlsx')
        print '[HuayClub] Save Excel Success!!'
        print '=============================='
    except:
        print "Something Error"


def writeExcel(sheet, role1, role2, value1, value2):
    print 'Hello'